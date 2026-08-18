"""Excel workbook generator for clinical prescriptions and medical bills."""

import io
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def generate_excel_export(
    structured_json: dict[str, Any],
    document_name: str = "Medical Document",
    document_id: str = "",
) -> bytes:
    """Generate a professionally styled XLSX workbook from structured medical document/bill data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Medical Document"

    # Styling Palette
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")  # Deep Navy Blue
    section_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")  # Slate 100
    table_header_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")  # Medical Evergreen/Teal
    total_fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")  # Mint Light
    accent_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # Amber highlight

    title_font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    section_font = Font(name="Segoe UI", size=11, bold=True, color="1E293B")
    label_font = Font(name="Segoe UI", size=10, bold=True, color="475569")
    value_font = Font(name="Segoe UI", size=10, color="0F172A")
    table_header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    total_font = Font(name="Segoe UI", size=11, bold=True, color="065F46")

    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    thick_bottom_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="medium", color="0F766E"),
    )

    # 1. Main Document Header Banner
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"MEDICAL DOCUMENT & BILL EXTRACTION REPORT: {document_name.upper()}"
    title_cell.font = title_font
    title_cell.fill = header_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Sub-header: Document Info & Export Time
    ws["A2"] = "Exported On:"
    ws["A2"].font = label_font
    ws["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["B2"].font = value_font

    ws["D2"] = "Document ID:"
    ws["D2"].font = label_font
    ws["E2"] = document_id or "N/A"
    ws["E2"].font = value_font

    current_row = 4

    # Extract Common Sections
    provider = structured_json.get("provider") or structured_json.get("clinician") or {}
    patient = structured_json.get("patient") or {}
    diagnosis = structured_json.get("diagnosis") or {}
    billing_summary = structured_json.get("billing_summary") or {}
    medicines = structured_json.get("medicines") or structured_json.get("items") or []

    # 2. Healthcare Provider & Hospital / Doctor Details
    if isinstance(provider, dict) and any(provider.values()):
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        sec_cell = ws.cell(row=current_row, column=1, value="1. HEALTHCARE PROVIDER & CLINICIAN DETAILS")
        sec_cell.font = section_font
        sec_cell.fill = section_fill
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        fields_to_show = [
            ("Hospital / Pharmacy", provider.get("hospital_name") or provider.get("name")),
            ("Doctor / Clinician", provider.get("doctor_name") or provider.get("doctor") or provider.get("specialty")),
            ("Bill / Invoice No", provider.get("bill_number") or provider.get("invoice_no")),
            ("Receipt / Bill Date", provider.get("bill_date") or provider.get("date")),
            ("Tax / GSTIN ID", provider.get("tax_id") or provider.get("gstin")),
            ("Contact Number", provider.get("contact_number") or provider.get("phone")),
        ]
        
        # Display in 2 columns
        for i in range(0, len(fields_to_show), 2):
            lbl1, val1 = fields_to_show[i]
            ws.cell(row=current_row, column=1, value=lbl1).font = label_font
            ws.cell(row=current_row, column=2, value=str(val1 or "—")).font = value_font
            
            if i + 1 < len(fields_to_show):
                lbl2, val2 = fields_to_show[i + 1]
                ws.cell(row=current_row, column=4, value=lbl2).font = label_font
                ws.cell(row=current_row, column=5, value=str(val2 or "—")).font = value_font
            current_row += 1

        current_row += 1

    # 3. Patient Details
    if isinstance(patient, dict) and any(patient.values()):
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        sec_cell = ws.cell(row=current_row, column=1, value="2. PATIENT & CUSTOMER INFORMATION")
        sec_cell.font = section_font
        sec_cell.fill = section_fill
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        p_fields = [
            ("Patient Name", patient.get("name") or patient.get("patient_name")),
            ("Patient ID / UHID", patient.get("patient_id") or patient.get("uhid")),
            ("Age", patient.get("age")),
            ("Gender", patient.get("gender")),
        ]
        for i in range(0, len(p_fields), 2):
            lbl1, val1 = p_fields[i]
            ws.cell(row=current_row, column=1, value=lbl1).font = label_font
            ws.cell(row=current_row, column=2, value=str(val1 or "—")).font = value_font
            if i + 1 < len(p_fields):
                lbl2, val2 = p_fields[i + 1]
                ws.cell(row=current_row, column=4, value=lbl2).font = label_font
                ws.cell(row=current_row, column=5, value=str(val2 or "—")).font = value_font
            current_row += 1

        current_row += 1

    # 4. Itemized Prescribed / Billed Medicines Table
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
    sec_cell = ws.cell(row=current_row, column=1, value="3. ITEMIZED MEDICINES, REMEDIES & BILLED ITEMS")
    sec_cell.font = section_font
    sec_cell.fill = section_fill
    ws.row_dimensions[current_row].height = 24
    current_row += 1

    headers = [
        "#",
        "Medicine / Item Name",
        "Unique Code / Batch",
        "Strength / Potency",
        "Unit Price / MRP",
        "Quantity",
        "Discount",
        "Total Price / Amount",
    ]

    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header_text)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = Alignment(horizontal="center" if col_idx in (1, 3, 4, 5, 6, 7, 8) else "left", vertical="center")
        cell.border = thin_border

    ws.row_dimensions[current_row].height = 22
    current_row += 1

    if isinstance(medicines, list) and medicines:
        for idx, item in enumerate(medicines, start=1):
            if not isinstance(item, dict):
                continue
            item_name = item.get("medicine_name") or item.get("item_name") or item.get("description") or item.get("name") or "—"
            unique_code = item.get("unique_code") or item.get("batch_no") or item.get("hsn_code") or item.get("code") or "—"
            strength = item.get("strength") or item.get("potency") or item.get("frequency") or "—"
            unit_price = item.get("unit_price") or item.get("price") or item.get("mrp") or item.get("rate")
            quantity = item.get("quantity") or item.get("qty")
            discount = item.get("discount") or item.get("disc") or item.get("tax_rate")
            total_price = item.get("total_price") or item.get("amount")

            row_values = [
                idx,
                item_name,
                unique_code,
                strength,
                unit_price if unit_price is not None else "—",
                quantity if quantity is not None else "—",
                discount if discount is not None else "—",
                total_price if total_price is not None else "—",
            ]

            for col_idx, val in enumerate(row_values, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                cell.font = value_font
                cell.border = thin_border
                align = "center" if col_idx in (1, 3, 4, 5, 6, 7, 8) else "left"
                cell.alignment = Alignment(horizontal=align, vertical="center")

            current_row += 1
    else:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        empty_cell = ws.cell(row=current_row, column=1, value="No itemized medicines or billing rows detected.")
        empty_cell.font = value_font
        empty_cell.alignment = Alignment(horizontal="center", vertical="center")
        empty_cell.border = thin_border
        current_row += 1

    current_row += 1

    # 5. Financial Summary & Totals
    if isinstance(billing_summary, dict) and any(billing_summary.values()):
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        sec_cell = ws.cell(row=current_row, column=1, value="4. FINANCIAL SUMMARY & BILLING TOTALS")
        sec_cell.font = section_font
        sec_cell.fill = section_fill
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        summary_rows = [
            ("Subtotal / Taxable Amount", billing_summary.get("subtotal")),
            ("Discount Total", billing_summary.get("discount_total")),
            ("Tax Amount (GST / VAT)", billing_summary.get("tax_amount")),
            ("GRAND TOTAL / NET PAYABLE", billing_summary.get("total_cost") or billing_summary.get("grand_total")),
            ("Payment Mode & Status", f"{billing_summary.get('payment_mode') or '—'} / {billing_summary.get('payment_status') or '—'}"),
        ]

        for lbl, val in summary_rows:
            is_grand_total = "GRAND TOTAL" in lbl
            lbl_cell = ws.cell(row=current_row, column=5, value=lbl)
            val_cell = ws.cell(row=current_row, column=8, value=val if val is not None else "—")
            
            lbl_cell.font = total_font if is_grand_total else label_font
            val_cell.font = total_font if is_grand_total else value_font
            
            if is_grand_total:
                lbl_cell.fill = total_fill
                val_cell.fill = total_fill
                lbl_cell.border = thick_bottom_border
                val_cell.border = thick_bottom_border
            else:
                lbl_cell.border = thin_border
                val_cell.border = thin_border

            lbl_cell.alignment = Alignment(horizontal="right", vertical="center")
            val_cell.alignment = Alignment(horizontal="center", vertical="center")
            current_row += 1

    # 6. Diagnosis / Clinical Notes (if available)
    if isinstance(diagnosis, dict) and any(diagnosis.values()):
        current_row += 1
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        sec_cell = ws.cell(row=current_row, column=1, value="5. CLINICAL DIAGNOSIS & INSTRUCTIONS")
        sec_cell.font = section_font
        sec_cell.fill = section_fill
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        for k, v in diagnosis.items():
            if v:
                ws.cell(row=current_row, column=1, value=k.replace("_", " ").title()).font = label_font
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=8)
                ws.cell(row=current_row, column=2, value=str(v)).font = value_font
                current_row += 1

    # Auto-fit Column Widths cleanly
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len and not cell.coordinate in ws.merged_cells:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    # Specific tweaks for readability
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 24

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
